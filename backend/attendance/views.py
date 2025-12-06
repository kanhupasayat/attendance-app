from rest_framework import status, generics, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from django.utils import timezone
from django.db.models import Sum, Count, Q
from datetime import datetime, timedelta
import csv
import pytz
from django.http import HttpResponse

from .models import Attendance, OfficeLocation, RegularizationRequest, WFHRequest, Shift, CompOff


def get_india_date():
    """Get current date in India timezone (IST)"""
    india_tz = pytz.timezone('Asia/Kolkata')
    return timezone.now().astimezone(india_tz).date()
from .serializers import (
    AttendanceSerializer, PunchInSerializer, PunchOutSerializer,
    OfficeLocationSerializer, AttendanceReportSerializer,
    RegularizationRequestSerializer, RegularizationApplySerializer,
    RegularizationReviewSerializer, WFHRequestSerializer,
    WFHApplySerializer, WFHReviewSerializer,
    AdminAttendanceCreateSerializer, AdminAttendanceUpdateSerializer,
    ShiftSerializer, ShiftCreateSerializer, CompOffSerializer, CompOffUseSerializer
)
from .utils import validate_location, validate_ip, get_client_ip
from accounts.views import IsAdminUser
from accounts.utils import (
    notify_regularization_applied, notify_regularization_status,
    notify_wfh_applied, notify_wfh_status
)
from accounts.activity_utils import (
    log_punch_in, log_punch_out, log_regularization_applied,
    log_regularization_reviewed, log_attendance_edit, log_wfh_applied,
    log_wfh_reviewed, log_shift_created, log_shift_updated, log_shift_deleted,
    log_holiday_added, log_holiday_updated, log_holiday_deleted
)


class PunchInView(APIView):
    def post(self, request):
        from leaves.models import LeaveRequest, LeaveBalance
        from django.db.models import Sum

        serializer = PunchInSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        latitude = serializer.validated_data.get('latitude')
        longitude = serializer.validated_data.get('longitude')
        face_verified = serializer.validated_data.get('face_verified', False)
        client_ip = get_client_ip(request)
        today = get_india_date()

        # Check if user is permanent WFH or has approved WFH for today
        is_wfh = request.user.is_permanent_wfh or WFHRequest.objects.filter(
            user=request.user,
            date=today,
            status='approved'
        ).exists()

        # Only validate location and IP if NOT WFH
        if not is_wfh:
            # Validate location
            location_valid, location_msg = validate_location(latitude, longitude)
            if not location_valid:
                return Response(
                    {"error": location_msg},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate IP
            ip_valid, ip_msg = validate_ip(client_ip)
            if not ip_valid:
                return Response(
                    {"error": ip_msg},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Check if already punched in today
        existing = Attendance.objects.filter(
            user=request.user, date=today
        ).first()

        if existing and existing.punch_in:
            return Response(
                {"error": "Already punched in today"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if today is employee's weekly off day
        # Python weekday(): Monday=0, Sunday=6 (matches our model)
        current_weekday = today.weekday()
        is_off_day = (current_weekday == request.user.weekly_off)

        # AUTO LEAVE CANCEL: Check if employee has approved leave for today
        leave_cancelled_msg = ""
        approved_leave = LeaveRequest.objects.filter(
            user=request.user,
            status='approved',
            start_date__lte=today,
            end_date__gte=today
        ).first()

        if approved_leave:
            # Employee has approved leave but came to office - handle leave adjustment
            leave_cancelled_msg = self.handle_leave_on_punch_in(approved_leave, today)

        if existing:
            # Update existing record
            existing.punch_in = timezone.now()
            existing.punch_in_latitude = latitude
            existing.punch_in_longitude = longitude
            existing.punch_in_ip = client_ip
            existing.is_off_day = is_off_day
            existing.is_wfh = is_wfh
            existing.face_verified = face_verified
            existing.save()
            attendance = existing
        else:
            # Create new record
            attendance = Attendance.objects.create(
                user=request.user,
                date=today,
                punch_in=timezone.now(),
                punch_in_latitude=latitude,
                punch_in_longitude=longitude,
                punch_in_ip=client_ip,
                is_off_day=is_off_day,
                is_wfh=is_wfh,
                face_verified=face_verified
            )

        # Log activity
        try:
            log_punch_in(request.user, attendance, request)
        except Exception:
            pass  # Don't fail punch in if logging fails

        msg = "Punch in successful (Work From Home)" if is_wfh else "Punch in successful"
        if leave_cancelled_msg:
            msg += f". {leave_cancelled_msg}"

        return Response({
            "message": msg,
            "data": AttendanceSerializer(attendance).data
        }, status=status.HTTP_201_CREATED)

    def handle_leave_on_punch_in(self, leave_request, today):
        """
        Handle leave adjustment when employee punches in during approved leave.

        Cases:
        1. Single day leave (start == end == today) -> Cancel entire leave
        2. First day of multi-day leave (start == today) -> Shrink leave (start = today + 1)
        3. Last day of multi-day leave (end == today) -> Shrink leave (end = today - 1)
        4. Middle day of multi-day leave -> Split into two leaves
        """
        from leaves.models import LeaveRequest, LeaveBalance
        from django.db.models import Sum
        from datetime import timedelta

        start_date = leave_request.start_date
        end_date = leave_request.end_date

        # Calculate original days
        original_total_days = float(leave_request.total_days)

        if start_date == end_date == today:
            # Case 1: Single day leave - cancel entirely
            leave_request.status = 'cancelled'
            leave_request.review_remarks = f"Auto-cancelled: Employee punched in on {today}"
            leave_request.save()

            # Restore leave balance
            self.restore_leave_balance(leave_request)
            return "Your 1 day leave has been auto-cancelled"

        elif start_date == today:
            # Case 2: First day of multi-day leave - shrink from start
            new_start = today + timedelta(days=1)
            new_total_days = (end_date - new_start).days + 1
            if leave_request.is_half_day:
                new_total_days = 0.5

            # Recalculate paid/comp-off/lop for reduced days
            self.recalculate_leave_days(leave_request, new_total_days, new_start, end_date)
            leave_request.start_date = new_start
            leave_request.review_remarks = f"Auto-adjusted: Employee punched in on {today}. Original: {start_date} to {end_date}"
            leave_request.save()

            return f"Leave adjusted: Now {new_start} to {end_date} ({new_total_days} days)"

        elif end_date == today:
            # Case 3: Last day of multi-day leave - shrink from end
            new_end = today - timedelta(days=1)
            new_total_days = (new_end - start_date).days + 1
            if leave_request.is_half_day:
                new_total_days = 0.5

            # Recalculate paid/comp-off/lop for reduced days
            self.recalculate_leave_days(leave_request, new_total_days, start_date, new_end)
            leave_request.end_date = new_end
            leave_request.review_remarks = f"Auto-adjusted: Employee punched in on {today}. Original: {start_date} to {end_date}"
            leave_request.save()

            return f"Leave adjusted: Now {start_date} to {new_end} ({new_total_days} days)"

        else:
            # Case 4: Middle day - split into two leaves
            # First part: start_date to (today - 1)
            first_end = today - timedelta(days=1)
            first_total_days = (first_end - start_date).days + 1

            # Second part: (today + 1) to end_date
            second_start = today + timedelta(days=1)
            second_total_days = (end_date - second_start).days + 1

            # Update original leave to first part
            self.recalculate_leave_days(leave_request, first_total_days, start_date, first_end)
            leave_request.end_date = first_end
            leave_request.review_remarks = f"Auto-split: Employee punched in on {today}. Original: {start_date} to {end_date}"
            leave_request.save()

            # Create new leave for second part
            if second_total_days > 0:
                # Calculate days for second leave
                second_leave = LeaveRequest.objects.create(
                    user=leave_request.user,
                    leave_type=leave_request.leave_type,
                    start_date=second_start,
                    end_date=end_date,
                    reason=leave_request.reason,
                    status='approved',
                    is_half_day=False,
                    total_days=second_total_days,
                    comp_off_days=0,
                    paid_days=0,
                    lop_days=0,
                    reviewed_by=leave_request.reviewed_by,
                    review_remarks=f"Auto-created from split: Original leave {start_date} to {end_date}"
                )
                # Recalculate days for second leave
                self.recalculate_leave_days(second_leave, second_total_days, second_start, end_date)
                second_leave.save()

                return f"Leave split: {start_date} to {first_end} ({first_total_days} days) + {second_start} to {end_date} ({second_total_days} days). {today} cancelled."

            return f"Leave adjusted: Now {start_date} to {first_end} ({first_total_days} days)"

    def recalculate_leave_days(self, leave_request, new_total_days, new_start, new_end):
        """
        Recalculate comp_off_days, paid_days, and lop_days for adjusted leave.
        Uses Smart Leave System: Comp Off first, then Paid, then LOP.
        """
        from leaves.models import LeaveBalance, LeaveRequest
        from django.db.models import Sum

        user = leave_request.user
        leave_type = leave_request.leave_type
        year = new_start.year
        month = new_start.month

        # Get available comp off (status='earned' means not used)
        available_comp_off = CompOff.objects.filter(
            user=user,
            status='earned',
            earned_date__lte=new_end
        ).count()

        # Get leave balance for this month
        balance = LeaveBalance.objects.filter(
            user=user,
            leave_type=leave_type,
            year=year,
            month=month
        ).first()

        available_paid = 0
        if balance:
            # Get already used paid days (excluding current leave being adjusted)
            already_used = LeaveRequest.objects.filter(
                user=user,
                leave_type=leave_type,
                status='approved',
                start_date__year=year,
                start_date__month=month
            ).exclude(id=leave_request.id).aggregate(total=Sum('paid_days'))['total'] or 0

            available_paid = float(balance.total_leaves) + float(balance.carried_forward) - float(already_used)
            available_paid = max(0, available_paid)

        remaining_days = float(new_total_days)

        # 1. Use Comp Off first
        comp_off_to_use = min(available_comp_off, remaining_days)
        remaining_days -= comp_off_to_use

        # 2. Use Paid leaves
        paid_to_use = min(available_paid, remaining_days)
        remaining_days -= paid_to_use

        # 3. Remaining is LOP
        lop_days = remaining_days

        # Update leave request
        old_comp_off = float(leave_request.comp_off_days or 0)

        leave_request.total_days = new_total_days
        leave_request.comp_off_days = comp_off_to_use
        leave_request.paid_days = paid_to_use
        leave_request.lop_days = lop_days

        # Restore excess comp offs that were used
        comp_off_to_restore = old_comp_off - comp_off_to_use
        if comp_off_to_restore > 0:
            # Mark some comp offs as earned (restore them)
            used_comp_offs = CompOff.objects.filter(
                user=user,
                status='used'
            ).order_by('-used_date')[:int(comp_off_to_restore)]
            for co in used_comp_offs:
                co.status = 'earned'
                co.used_date = None
                co.save()

    def restore_leave_balance(self, leave_request):
        """Restore leave balance when leave is cancelled"""
        # Restore comp offs used in this leave
        comp_off_to_restore = int(float(leave_request.comp_off_days or 0))
        if comp_off_to_restore > 0:
            used_comp_offs = CompOff.objects.filter(
                user=leave_request.user,
                status='used'
            ).order_by('-used_date')[:comp_off_to_restore]
            for co in used_comp_offs:
                co.status = 'earned'
                co.used_date = None
                co.save()


class PunchOutView(APIView):
    def post(self, request):
        serializer = PunchOutSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        latitude = serializer.validated_data.get('latitude')
        longitude = serializer.validated_data.get('longitude')
        client_ip = get_client_ip(request)
        today = get_india_date()

        # Check if user is permanent WFH or has approved WFH for today
        is_wfh = request.user.is_permanent_wfh or WFHRequest.objects.filter(
            user=request.user,
            date=today,
            status='approved'
        ).exists()

        # Only validate location and IP if NOT WFH
        if not is_wfh:
            # Validate location
            location_valid, location_msg = validate_location(latitude, longitude)
            if not location_valid:
                return Response(
                    {"error": location_msg},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate IP
            ip_valid, ip_msg = validate_ip(client_ip)
            if not ip_valid:
                return Response(
                    {"error": ip_msg},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Check if punched in today
        attendance = Attendance.objects.filter(
            user=request.user, date=today
        ).first()

        if not attendance or not attendance.punch_in:
            return Response(
                {"error": "You haven't punched in today"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if attendance.punch_out:
            return Response(
                {"error": "Already punched out today"},
                status=status.HTTP_400_BAD_REQUEST
            )

        attendance.punch_out = timezone.now()
        attendance.punch_out_latitude = latitude
        attendance.punch_out_longitude = longitude
        attendance.punch_out_ip = client_ip
        attendance.save()  # This will calculate working hours

        # Log activity
        try:
            log_punch_out(request.user, attendance, request)
        except Exception:
            pass  # Don't fail punch out if logging fails

        msg = "Punch out successful (Work From Home)" if is_wfh else "Punch out successful"
        return Response({
            "message": msg,
            "data": AttendanceSerializer(attendance).data
        })


class TodayAttendanceView(APIView):
    def get(self, request):
        today = get_india_date()
        attendance = Attendance.objects.filter(
            user=request.user, date=today
        ).first()

        if attendance:
            return Response(AttendanceSerializer(attendance).data)
        return Response({"message": "No attendance record for today"})


class MyAttendanceListView(generics.ListAPIView):
    serializer_class = AttendanceSerializer

    def get_queryset(self):
        queryset = Attendance.objects.filter(
            user=self.request.user
        ).select_related('user')

        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')

        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])
        elif month and year:
            queryset = queryset.filter(
                date__month=month, date__year=year
            )

        return queryset.order_by('-date')


# Admin views
class AllAttendanceListView(generics.ListAPIView):
    permission_classes = [IsAdminUser]
    serializer_class = AttendanceSerializer

    def get_queryset(self):
        queryset = Attendance.objects.select_related('user')

        user_id = self.request.query_params.get('user_id')
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')

        if user_id:
            queryset = queryset.filter(user_id=user_id)
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])
        elif month and year:
            queryset = queryset.filter(date__month=month, date__year=year)

        return queryset.order_by('-date', 'user__name')


class AttendanceReportView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        month = int(request.query_params.get('month', timezone.now().month))
        year = int(request.query_params.get('year', timezone.now().year))

        from accounts.models import User
        from django.db.models import Value, DecimalField
        from django.db.models.functions import Coalesce
        from decimal import Decimal

        # Single optimized query with annotation - eliminates N+1 problem
        report = User.objects.filter(
            role='employee', is_active=True
        ).annotate(
            total_present=Count(
                'attendances',
                filter=Q(attendances__date__month=month, attendances__date__year=year, attendances__status='present')
            ),
            total_absent=Count(
                'attendances',
                filter=Q(attendances__date__month=month, attendances__date__year=year, attendances__status='absent')
            ),
            total_half_day=Count(
                'attendances',
                filter=Q(attendances__date__month=month, attendances__date__year=year, attendances__status='half_day')
            ),
            total_on_leave=Count(
                'attendances',
                filter=Q(attendances__date__month=month, attendances__date__year=year, attendances__status='on_leave')
            ),
            total_working_hours=Coalesce(
                Sum('attendances__working_hours',
                    filter=Q(attendances__date__month=month, attendances__date__year=year)),
                Value(Decimal('0'), output_field=DecimalField())
            )
        ).values(
            'id', 'name', 'total_present', 'total_absent',
            'total_half_day', 'total_on_leave', 'total_working_hours'
        ).order_by('name')

        # Format response
        result = [{
            'user_id': emp['id'],
            'user_name': emp['name'],
            'total_present': emp['total_present'],
            'total_absent': emp['total_absent'],
            'total_half_day': emp['total_half_day'],
            'total_on_leave': emp['total_on_leave'],
            'total_working_hours': float(emp['total_working_hours']) if emp['total_working_hours'] else 0
        } for emp in report]

        return Response(result)


class ExportAttendanceCSVView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        month = int(request.query_params.get('month', timezone.now().month))
        year = int(request.query_params.get('year', timezone.now().year))

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="attendance_{year}_{month}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Employee Name', 'Date', 'Punch In', 'Punch Out',
            'Working Hours', 'Status'
        ])

        attendances = Attendance.objects.filter(
            date__month=month, date__year=year
        ).select_related('user').order_by('user__name', 'date')

        for att in attendances:
            writer.writerow([
                att.user.name,
                att.date,
                att.punch_in.strftime('%H:%M:%S') if att.punch_in else '',
                att.punch_out.strftime('%H:%M:%S') if att.punch_out else '',
                att.working_hours,
                att.status
            ])

        return response


class ExportAttendanceExcelView(APIView):
    """Export attendance data to Excel format"""
    permission_classes = [IsAdminUser]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO

        month = int(request.query_params.get('month', timezone.now().month))
        year = int(request.query_params.get('year', timezone.now().year))

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance {year}-{month:02d}"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = ['Employee Name', 'Date', 'Punch In', 'Punch Out', 'Working Hours', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data
        attendances = Attendance.objects.filter(
            date__month=month, date__year=year
        ).select_related('user').order_by('user__name', 'date')

        for row, att in enumerate(attendances, 2):
            ws.cell(row=row, column=1, value=att.user.name).border = thin_border
            ws.cell(row=row, column=2, value=str(att.date)).border = thin_border
            ws.cell(row=row, column=3, value=att.punch_in.strftime('%I:%M %p') if att.punch_in else '-').border = thin_border
            ws.cell(row=row, column=4, value=att.punch_out.strftime('%I:%M %p') if att.punch_out else '-').border = thin_border
            ws.cell(row=row, column=5, value=float(att.working_hours) if att.working_hours else 0).border = thin_border
            ws.cell(row=row, column=6, value=att.status.upper()).border = thin_border

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12

        # Save to response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="attendance_{year}_{month:02d}.xlsx"'
        return response


class ExportAttendancePDFView(APIView):
    """Export attendance data to PDF format"""
    permission_classes = [IsAdminUser]

    def get(self, request):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from io import BytesIO

        month = int(request.query_params.get('month', timezone.now().month))
        year = int(request.query_params.get('year', timezone.now().year))

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,
            spaceAfter=20
        )

        # Title
        month_name = ['January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December'][month - 1]
        title = Paragraph(f"Attendance Report - {month_name} {year}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))

        # Data
        attendances = Attendance.objects.filter(
            date__month=month, date__year=year
        ).select_related('user').order_by('user__name', 'date')

        data = [['Employee', 'Date', 'Punch In', 'Punch Out', 'Hours', 'Status']]
        for att in attendances:
            data.append([
                att.user.name,
                str(att.date),
                att.punch_in.strftime('%I:%M %p') if att.punch_in else '-',
                att.punch_out.strftime('%I:%M %p') if att.punch_out else '-',
                str(att.working_hours) if att.working_hours else '0',
                att.status.upper()
            ])

        if len(data) > 1:
            table = Table(data, colWidths=[120, 80, 80, 80, 60, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F2F2F2')),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("No attendance records found for this period.", styles['Normal']))

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="attendance_{year}_{month:02d}.pdf"'
        return response


class OfficeLocationListView(generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]
    serializer_class = OfficeLocationSerializer
    queryset = OfficeLocation.objects.all()


class OfficeLocationDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]
    serializer_class = OfficeLocationSerializer
    queryset = OfficeLocation.objects.all()


class OffDayWorkStatsView(APIView):
    """Get off-day work stats for the current user"""

    def get(self, request):
        month = int(request.query_params.get('month', timezone.now().month))
        year = int(request.query_params.get('year', timezone.now().year))

        # Get all off-day attendance records for this user
        off_day_records = Attendance.objects.filter(
            user=request.user,
            is_off_day=True,
            date__month=month,
            date__year=year
        ).order_by('-date')

        # Count total off-day work days
        total_off_day_work = off_day_records.count()

        # Calculate total hours worked on off days
        total_hours = off_day_records.aggregate(
            total=Sum('working_hours')
        )['total'] or 0

        # Get the list of dates worked on off days
        off_day_dates = list(off_day_records.values('date', 'working_hours'))

        return Response({
            'total_off_day_work': total_off_day_work,
            'total_hours_on_off_days': total_hours,
            'off_day_records': off_day_dates,
            'month': month,
            'year': year
        })


# Regularization Views
class RegularizationApplyView(APIView):
    """Employee applies for attendance regularization"""

    def post(self, request):
        serializer = RegularizationApplySerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        date = serializer.validated_data['date']
        request_type = serializer.validated_data['request_type']
        requested_punch_in = serializer.validated_data.get('requested_punch_in')
        requested_punch_out = serializer.validated_data.get('requested_punch_out')
        reason = serializer.validated_data['reason']

        # Cannot apply regularization for future dates (today is allowed)
        today = get_india_date()
        if date > today:
            return Response(
                {"error": "Regularization cannot be applied for future dates"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if request already exists for this date
        existing = RegularizationRequest.objects.filter(
            user=request.user,
            date=date,
            status='pending'
        ).first()

        if existing:
            return Response(
                {"error": "A pending regularization request already exists for this date"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get existing attendance record if any
        attendance = Attendance.objects.filter(
            user=request.user,
            date=date
        ).first()

        # Create regularization request
        regularization = RegularizationRequest.objects.create(
            user=request.user,
            attendance=attendance,
            date=date,
            request_type=request_type,
            requested_punch_in=requested_punch_in,
            requested_punch_out=requested_punch_out,
            reason=reason
        )

        # Notify admins about new regularization request
        notify_regularization_applied(regularization)

        # Log activity
        try:
            log_regularization_applied(request.user, regularization, request)
        except Exception:
            pass

        return Response({
            "message": "Regularization request submitted successfully",
            "data": RegularizationRequestSerializer(regularization).data
        }, status=status.HTTP_201_CREATED)


class MyRegularizationListView(generics.ListAPIView):
    """Employee views their own regularization requests"""
    serializer_class = RegularizationRequestSerializer

    def get_queryset(self):
        return RegularizationRequest.objects.filter(user=self.request.user)


class AllRegularizationListView(generics.ListAPIView):
    """Admin views all regularization requests"""
    permission_classes = [IsAdminUser]
    serializer_class = RegularizationRequestSerializer

    def get_queryset(self):
        queryset = RegularizationRequest.objects.all()

        status_filter = self.request.query_params.get('status')
        user_id = self.request.query_params.get('user_id')

        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if user_id:
            queryset = queryset.filter(user_id=user_id)

        return queryset


class RegularizationReviewView(APIView):
    """Admin reviews (approves/rejects) a regularization request"""
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        try:
            regularization = RegularizationRequest.objects.get(pk=pk)
        except RegularizationRequest.DoesNotExist:
            return Response(
                {"error": "Regularization request not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        if regularization.status != 'pending':
            return Response(
                {"error": "This request has already been reviewed"},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = RegularizationReviewSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        new_status = serializer.validated_data['status']
        review_remarks = serializer.validated_data.get('review_remarks', '')

        # VALIDATION: Check if requested time is in the future (only for approval)
        if new_status == 'approved':
            now = timezone.now()
            today = get_india_date()

            # If regularization is for today, check if requested times are in the future
            if regularization.date == today:
                current_time = now.astimezone(pytz.timezone('Asia/Kolkata')).time()

                if regularization.requested_punch_in:
                    if regularization.requested_punch_in > current_time:
                        return Response(
                            {"error": f"Cannot set punch in time ({regularization.requested_punch_in.strftime('%I:%M %p')}) in the future. Current time is {current_time.strftime('%I:%M %p')}. Please wait until that time has passed."},
                            status=status.HTTP_400_BAD_REQUEST
                        )

                if regularization.requested_punch_out:
                    if regularization.requested_punch_out > current_time:
                        return Response(
                            {"error": f"Cannot set punch out time ({regularization.requested_punch_out.strftime('%I:%M %p')}) in the future. Current time is {current_time.strftime('%I:%M %p')}. Please wait until that time has passed."},
                            status=status.HTTP_400_BAD_REQUEST
                        )

        regularization.status = new_status
        regularization.reviewed_by = request.user
        regularization.reviewed_on = timezone.now()
        regularization.review_remarks = review_remarks
        regularization.save()

        # If approved, update the attendance record
        if new_status == 'approved':
            attendance, created = Attendance.objects.get_or_create(
                user=regularization.user,
                date=regularization.date,
                defaults={'status': 'present'}
            )

            # Update attendance based on request type
            if regularization.requested_punch_in:
                punch_in_datetime = timezone.make_aware(
                    datetime.combine(regularization.date, regularization.requested_punch_in)
                )
                attendance.punch_in = punch_in_datetime

            if regularization.requested_punch_out:
                punch_out_datetime = timezone.make_aware(
                    datetime.combine(regularization.date, regularization.requested_punch_out)
                )
                attendance.punch_out = punch_out_datetime

            # Always set status to present when regularization is approved
            attendance.status = 'present'
            attendance.notes = f"Regularized: {regularization.request_type}"
            attendance.save(force_status=True)  # Prevent auto-status calculation

        # Notify employee about regularization status (with email)
        notify_regularization_status(regularization, new_status, review_remarks)

        # Log activity
        try:
            log_regularization_reviewed(request.user, regularization, new_status, request)
        except Exception:
            pass

        return Response({
            "message": f"Regularization request {new_status}",
            "data": RegularizationRequestSerializer(regularization).data
        })


class CancelRegularizationView(APIView):
    """Employee cancels their own pending regularization request"""

    def post(self, request, pk):
        try:
            regularization = RegularizationRequest.objects.get(
                pk=pk,
                user=request.user
            )
        except RegularizationRequest.DoesNotExist:
            return Response(
                {"error": "Regularization request not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        if regularization.status != 'pending':
            return Response(
                {"error": "Only pending requests can be cancelled"},
                status=status.HTTP_400_BAD_REQUEST
            )

        regularization.delete()

        return Response({"message": "Regularization request cancelled"})


# WFH (Work From Home) Views
class WFHApplyView(APIView):
    """Employee applies for Work From Home (supports date range)"""

    def post(self, request):
        serializer = WFHApplySerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        start_date = serializer.validated_data['start_date']
        end_date = serializer.validated_data.get('end_date') or start_date
        reason = serializer.validated_data['reason']
        today = get_india_date()

        # Cannot apply WFH for past dates
        if start_date < today:
            return Response(
                {"error": "WFH can only be applied for today or future dates"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Generate list of dates
        from datetime import timedelta
        dates = []
        current_date = start_date
        while current_date <= end_date:
            dates.append(current_date)
            current_date += timedelta(days=1)

        # Check for existing requests
        existing_dates = []
        for date in dates:
            existing = WFHRequest.objects.filter(
                user=request.user,
                date=date
            ).first()

            if existing:
                if existing.status == 'pending':
                    existing_dates.append(str(date))
                elif existing.status == 'approved':
                    existing_dates.append(f"{date} (approved)")
                elif existing.status == 'rejected':
                    # Allow reapplying if previously rejected
                    existing.delete()

        if existing_dates:
            return Response(
                {"error": f"WFH request already exists for: {', '.join(existing_dates)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create WFH requests for all dates
        created_requests = []
        for date in dates:
            wfh_request = WFHRequest.objects.create(
                user=request.user,
                date=date,
                reason=reason
            )
            created_requests.append(wfh_request)

        # Notify admins about new WFH request (only once for the first request)
        if created_requests:
            notify_wfh_applied(created_requests[0])

        # Log activity
        try:
            log_wfh_applied(request.user, created_requests[0], request)
        except Exception:
            pass

        # Prepare response message
        if len(dates) == 1:
            msg = f"WFH request submitted for {start_date}"
        else:
            msg = f"WFH requests submitted for {len(dates)} days ({start_date} to {end_date})"

        return Response({
            "message": msg,
            "data": WFHRequestSerializer(created_requests, many=True).data
        }, status=status.HTTP_201_CREATED)


class MyWFHListView(generics.ListAPIView):
    """Employee views their own WFH requests"""
    serializer_class = WFHRequestSerializer

    def get_queryset(self):
        return WFHRequest.objects.filter(user=self.request.user)


class AllWFHListView(generics.ListAPIView):
    """Admin views all WFH requests"""
    permission_classes = [IsAdminUser]
    serializer_class = WFHRequestSerializer

    def get_queryset(self):
        queryset = WFHRequest.objects.all()

        status_filter = self.request.query_params.get('status')
        user_id = self.request.query_params.get('user_id')

        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if user_id:
            queryset = queryset.filter(user_id=user_id)

        return queryset


class WFHReviewView(APIView):
    """Admin reviews (approves/rejects) a WFH request"""
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        try:
            wfh_request = WFHRequest.objects.get(pk=pk)
        except WFHRequest.DoesNotExist:
            return Response(
                {"error": "WFH request not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        if wfh_request.status != 'pending':
            return Response(
                {"error": "This request has already been reviewed"},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = WFHReviewSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        new_status = serializer.validated_data['status']
        review_remarks = serializer.validated_data.get('review_remarks', '')

        wfh_request.status = new_status
        wfh_request.reviewed_by = request.user
        wfh_request.reviewed_on = timezone.now()
        wfh_request.review_remarks = review_remarks
        wfh_request.save()

        # Notify employee about WFH status
        notify_wfh_status(wfh_request, new_status, review_remarks)

        # Log activity
        try:
            log_wfh_reviewed(request.user, wfh_request, new_status, request)
        except Exception:
            pass

        return Response({
            "message": f"WFH request {new_status}",
            "data": WFHRequestSerializer(wfh_request).data
        })


class BulkWFHReviewView(APIView):
    """Admin can approve/reject multiple WFH requests at once"""
    permission_classes = [IsAdminUser]

    def post(self, request):
        wfh_ids = request.data.get('wfh_ids', [])
        new_status = request.data.get('status')
        review_remarks = request.data.get('review_remarks', '')

        if not wfh_ids:
            return Response(
                {"error": "wfh_ids is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if new_status not in ['approved', 'rejected']:
            return Response(
                {"error": "status must be 'approved' or 'rejected'"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get all pending WFH requests
        wfh_requests = WFHRequest.objects.filter(
            id__in=wfh_ids,
            status='pending'
        )

        updated_count = 0
        for wfh_request in wfh_requests:
            wfh_request.status = new_status
            wfh_request.reviewed_by = request.user
            wfh_request.reviewed_on = timezone.now()
            wfh_request.review_remarks = review_remarks
            wfh_request.save()

            # Notify employee
            notify_wfh_status(wfh_request, new_status, review_remarks)

            # Log activity
            try:
                log_wfh_reviewed(request.user, wfh_request, new_status, request)
            except Exception:
                pass

            updated_count += 1

        return Response({
            "message": f"{updated_count} WFH request(s) {new_status}",
            "updated_count": updated_count
        })


class CancelWFHView(APIView):
    """Employee cancels their own pending WFH request"""

    def post(self, request, pk):
        try:
            wfh_request = WFHRequest.objects.get(
                pk=pk,
                user=request.user
            )
        except WFHRequest.DoesNotExist:
            return Response(
                {"error": "WFH request not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        if wfh_request.status != 'pending':
            return Response(
                {"error": "Only pending requests can be cancelled"},
                status=status.HTTP_400_BAD_REQUEST
            )

        wfh_request.delete()

        return Response({"message": "WFH request cancelled"})


class TodayWFHStatusView(APIView):
    """Check if user has approved WFH for today"""

    def get(self, request):
        today = get_india_date()
        wfh_request = WFHRequest.objects.filter(
            user=request.user,
            date=today,
            status='approved'
        ).first()

        return Response({
            "is_wfh": wfh_request is not None,
            "wfh_request": WFHRequestSerializer(wfh_request).data if wfh_request else None
        })


class AutoPunchOutView(APIView):
    """
    API endpoint to trigger auto punch-out for all employees who forgot to punch out.
    Secured with CRON_SECRET_KEY.
    Can be called by external cron services like cron-job.org
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        from django.conf import settings
        from .email_utils import send_auto_punch_out_email

        # Verify secret key
        secret_key = request.headers.get('X-Cron-Secret') or request.data.get('secret_key')
        expected_key = getattr(settings, 'CRON_SECRET_KEY', 'attendance-auto-punch-2024')

        if secret_key != expected_key:
            return Response(
                {"error": "Unauthorized"},
                status=status.HTTP_401_UNAUTHORIZED
            )

        today = get_india_date()
        now = timezone.now()

        # Find all attendance records with punch_in but no punch_out
        pending_punch_outs = Attendance.objects.filter(
            date=today,
            punch_in__isnull=False,
            punch_out__isnull=True
        )

        count = 0
        punched_out_users = []

        for attendance in pending_punch_outs:
            # Auto punch out at current time
            attendance.punch_out = now
            attendance.is_auto_punch_out = True
            attendance.notes = f"Auto punch out by system. Employee forgot to punch out."
            attendance.save()

            punched_out_users.append(attendance.user.name)

            # Send warning email to employee
            try:
                send_auto_punch_out_email(attendance)
            except Exception as e:
                print(f"Failed to send email to {attendance.user.name}: {e}")

            count += 1

        return Response({
            "message": f"Auto punch out completed for {count} employee(s)",
            "employees": punched_out_users
        })


# Admin Attendance Management Views
class AdminAttendanceCreateView(APIView):
    """Admin can add attendance for any employee"""
    permission_classes = [IsAdminUser]

    def post(self, request):
        serializer = AdminAttendanceCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        from accounts.models import User

        user_id = serializer.validated_data['user_id']
        date = serializer.validated_data['date']
        punch_in_time = serializer.validated_data.get('punch_in')
        punch_out_time = serializer.validated_data.get('punch_out')
        attendance_status = serializer.validated_data.get('status', 'present')
        notes = serializer.validated_data.get('notes', '')

        # Verify user exists
        try:
            user = User.objects.get(pk=user_id, role='employee')
        except User.DoesNotExist:
            return Response(
                {"error": "Employee not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if attendance already exists for this date
        existing = Attendance.objects.filter(user=user, date=date).first()
        if existing:
            return Response(
                {"error": f"Attendance already exists for {user.name} on {date}. Use update instead."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create attendance record
        attendance = Attendance(
            user=user,
            date=date,
            status=attendance_status,
            notes=f"Added by admin: {notes}" if notes else "Added by admin"
        )

        # Set punch times if provided
        if punch_in_time:
            attendance.punch_in = timezone.make_aware(
                datetime.combine(date, punch_in_time)
            )

        if punch_out_time:
            attendance.punch_out = timezone.make_aware(
                datetime.combine(date, punch_out_time)
            )

        attendance.save()

        return Response({
            "message": f"Attendance added for {user.name} on {date}",
            "data": AttendanceSerializer(attendance).data
        }, status=status.HTTP_201_CREATED)


class AdminAttendanceUpdateView(APIView):
    """Admin can update any attendance record"""
    permission_classes = [IsAdminUser]

    def patch(self, request, pk):
        try:
            attendance = Attendance.objects.get(pk=pk)
        except Attendance.DoesNotExist:
            return Response(
                {"error": "Attendance record not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = AdminAttendanceUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Update fields if provided
        if 'punch_in' in request.data:
            punch_in_time = serializer.validated_data.get('punch_in')
            if punch_in_time:
                attendance.punch_in = timezone.make_aware(
                    datetime.combine(attendance.date, punch_in_time)
                )
            else:
                attendance.punch_in = None

        if 'punch_out' in request.data:
            punch_out_time = serializer.validated_data.get('punch_out')
            if punch_out_time:
                attendance.punch_out = timezone.make_aware(
                    datetime.combine(attendance.date, punch_out_time)
                )
            else:
                attendance.punch_out = None

        # Check if admin is explicitly setting status
        force_status = False
        if 'status' in serializer.validated_data:
            attendance.status = serializer.validated_data['status']
            force_status = True  # Preserve admin's status choice

        if 'notes' in serializer.validated_data:
            existing_notes = attendance.notes or ''
            new_note = serializer.validated_data['notes']
            attendance.notes = f"{existing_notes}\n[Admin update]: {new_note}".strip()

        attendance.save(force_status=force_status)

        # Log activity
        try:
            changes = {k: v for k, v in request.data.items() if k != 'notes'}
            log_attendance_edit(request.user, attendance, changes, request)
        except Exception:
            pass

        return Response({
            "message": "Attendance updated successfully",
            "data": AttendanceSerializer(attendance).data
        })


class AdminMarkAbsentView(APIView):
    """Admin can mark employee as absent for a specific date"""
    permission_classes = [IsAdminUser]

    def post(self, request):
        from accounts.models import User

        user_id = request.data.get('user_id')
        date_str = request.data.get('date')
        notes = request.data.get('notes', '')

        if not user_id or not date_str:
            return Response(
                {"error": "user_id and date are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = User.objects.get(pk=user_id, role='employee')
        except User.DoesNotExist:
            return Response(
                {"error": "Employee not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Create or update attendance record
        attendance, created = Attendance.objects.update_or_create(
            user=user,
            date=date,
            defaults={
                'status': 'absent',
                'punch_in': None,
                'punch_out': None,
                'working_hours': 0,
                'notes': f"Marked absent by admin: {notes}" if notes else "Marked absent by admin"
            }
        )

        action = "marked" if created else "updated to"
        return Response({
            "message": f"{user.name} {action} absent on {date}",
            "data": AttendanceSerializer(attendance).data
        })


class AdminBulkAttendanceView(APIView):
    """Admin can mark multiple employees absent/present for a date"""
    permission_classes = [IsAdminUser]

    def post(self, request):
        from accounts.models import User

        date_str = request.data.get('date')
        user_ids = request.data.get('user_ids', [])
        attendance_status = request.data.get('status', 'absent')
        notes = request.data.get('notes', '')

        if not date_str or not user_ids:
            return Response(
                {"error": "date and user_ids are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if attendance_status not in ['present', 'absent', 'half_day', 'on_leave']:
            return Response(
                {"error": "Invalid status. Use: present, absent, half_day, on_leave"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST
            )

        results = []
        for user_id in user_ids:
            try:
                user = User.objects.get(pk=user_id, role='employee')
                attendance, created = Attendance.objects.update_or_create(
                    user=user,
                    date=date,
                    defaults={
                        'status': attendance_status,
                        'notes': f"Bulk update by admin: {notes}" if notes else "Bulk update by admin"
                    }
                )
                results.append({
                    'user_id': user_id,
                    'user_name': user.name,
                    'status': 'success'
                })
            except User.DoesNotExist:
                results.append({
                    'user_id': user_id,
                    'status': 'failed',
                    'error': 'Employee not found'
                })

        return Response({
            "message": f"Bulk attendance update completed for {date}",
            "results": results
        })


class AdminClearPunchOutView(APIView):
    """Admin can clear punch out so employee can punch out again"""
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        try:
            attendance = Attendance.objects.get(pk=pk)
        except Attendance.DoesNotExist:
            return Response(
                {"error": "Attendance record not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        if not attendance.punch_out:
            return Response(
                {"error": "No punch out to clear"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Store old punch out for notes
        old_punch_out = attendance.punch_out

        # Clear punch out
        attendance.punch_out = None
        attendance.punch_out_latitude = None
        attendance.punch_out_longitude = None
        attendance.punch_out_ip = None
        attendance.working_hours = 0
        attendance.status = 'present'  # Reset to present (will update on new punch out)

        # Add note about clearing
        existing_notes = attendance.notes or ''
        attendance.notes = f"{existing_notes}\n[Admin cleared punch out: {old_punch_out.strftime('%H:%M')}]".strip()

        # Save without auto-calculating status
        attendance.save(force_status=True)

        return Response({
            "message": f"Punch out cleared for {attendance.user.name}. Employee can now punch out again.",
            "data": AttendanceSerializer(attendance).data
        })


# Shift Management Views
class ShiftListCreateView(generics.ListCreateAPIView):
    """List all shifts or create a new shift (Admin only)"""
    permission_classes = [IsAdminUser]
    queryset = Shift.objects.all()

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return ShiftCreateSerializer
        return ShiftSerializer

    def perform_create(self, serializer):
        shift = serializer.save()
        # Log activity
        try:
            log_shift_created(self.request.user, shift, self.request)
        except Exception:
            pass


class ShiftDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Get, update or delete a shift (Admin only)"""
    permission_classes = [IsAdminUser]
    serializer_class = ShiftSerializer
    queryset = Shift.objects.all()

    def perform_update(self, serializer):
        shift = serializer.save()
        try:
            log_shift_updated(self.request.user, shift, self.request)
        except Exception:
            pass

    def perform_destroy(self, instance):
        shift_name = instance.name
        instance.delete()
        try:
            log_shift_deleted(self.request.user, shift_name, self.request)
        except Exception:
            pass


class AssignShiftView(APIView):
    """Assign shift to employee(s)"""
    permission_classes = [IsAdminUser]

    def post(self, request):
        from accounts.models import User

        shift_id = request.data.get('shift_id')
        user_ids = request.data.get('user_ids', [])

        if not user_ids:
            return Response(
                {"error": "user_ids is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # shift_id can be null to remove shift
        shift = None
        if shift_id:
            try:
                shift = Shift.objects.get(pk=shift_id, is_active=True)
            except Shift.DoesNotExist:
                return Response(
                    {"error": "Shift not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

        updated_count = User.objects.filter(
            pk__in=user_ids,
            role='employee'
        ).update(shift=shift)

        shift_name = shift.name if shift else "No Shift"
        return Response({
            "message": f"Shift '{shift_name}' assigned to {updated_count} employee(s)",
            "updated_count": updated_count
        })


# Comp Off Views
class MyCompOffListView(generics.ListAPIView):
    """List current user's comp offs"""
    serializer_class = CompOffSerializer

    def get_queryset(self):
        return CompOff.objects.filter(user=self.request.user)


class AllCompOffListView(generics.ListAPIView):
    """List all comp offs (Admin only)"""
    permission_classes = [IsAdminUser]
    serializer_class = CompOffSerializer

    def get_queryset(self):
        queryset = CompOff.objects.all()
        status_filter = self.request.query_params.get('status', '')
        user_id = self.request.query_params.get('user_id', '')

        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if user_id:
            queryset = queryset.filter(user_id=user_id)

        return queryset


class CompOffBalanceView(APIView):
    """Get comp off balance for current user or specific user"""

    def get(self, request):
        user_id = request.query_params.get('user_id')

        if user_id and request.user.is_admin:
            from accounts.models import User
            try:
                user = User.objects.get(pk=user_id)
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found"},
                    status=status.HTTP_404_NOT_FOUND
                )
        else:
            user = request.user

        # Calculate balances
        earned = CompOff.objects.filter(user=user, status='earned').aggregate(
            total=Sum('credit_days')
        )['total'] or 0

        used = CompOff.objects.filter(user=user, status='used').aggregate(
            total=Sum('credit_days')
        )['total'] or 0

        expired = CompOff.objects.filter(user=user, status='expired').aggregate(
            total=Sum('credit_days')
        )['total'] or 0

        # Check and mark expired comp offs
        today = get_india_date()
        CompOff.objects.filter(
            user=user,
            status='earned',
            expires_on__lt=today
        ).update(status='expired')

        # Calculate pending comp off days in leave requests
        from leaves.models import LeaveRequest
        pending_in_leaves = LeaveRequest.objects.filter(
            user=user,
            status='pending'
        ).aggregate(total=Sum('comp_off_days'))['total'] or 0

        return Response({
            "user_id": user.id,
            "user_name": user.name,
            "earned": float(earned),
            "used": float(used),
            "expired": float(expired),
            "available": float(earned),  # Available = earned (status='earned')
            "pending_in_leaves": float(pending_in_leaves)  # Comp offs reserved in pending leaves
        })


class UseCompOffView(APIView):
    """Use comp off as leave"""

    def post(self, request):
        serializer = CompOffUseSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        comp_off_id = serializer.validated_data['comp_off_id']
        use_date = serializer.validated_data['use_date']

        try:
            comp_off = CompOff.objects.get(
                pk=comp_off_id,
                user=request.user,
                status='earned'
            )
        except CompOff.DoesNotExist:
            return Response(
                {"error": "Comp off not found or already used"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if expired
        if comp_off.expires_on and use_date > comp_off.expires_on:
            return Response(
                {"error": "Comp off has expired"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Mark comp off as used
        comp_off.status = 'used'
        comp_off.used_date = use_date
        comp_off.save()

        # Create attendance record for the comp off day
        attendance, created = Attendance.objects.update_or_create(
            user=request.user,
            date=use_date,
            defaults={
                'status': 'on_leave',
                'notes': f"Comp Off used (earned on {comp_off.earned_date})"
            }
        )

        return Response({
            "message": f"Comp off used successfully for {use_date}",
            "comp_off": CompOffSerializer(comp_off).data
        })


class AdminCreateCompOffView(APIView):
    """Admin can manually create comp off for employee"""
    permission_classes = [IsAdminUser]

    def post(self, request):
        from accounts.models import User

        user_id = request.data.get('user_id')
        earned_date = request.data.get('earned_date')
        credit_days = request.data.get('credit_days', 1.0)
        reason = request.data.get('reason', 'Manual credit by admin')

        if not user_id or not earned_date:
            return Response(
                {"error": "user_id and earned_date are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = User.objects.get(pk=user_id, role='employee')
        except User.DoesNotExist:
            return Response(
                {"error": "Employee not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            earned_date = datetime.strptime(earned_date, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if comp off already exists for this user and date
        existing = CompOff.objects.filter(
            user=user,
            earned_date=earned_date
        ).first()

        if existing:
            return Response(
                {"error": f"Comp off already exists for {user.name} on {earned_date}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        comp_off = CompOff.objects.create(
            user=user,
            earned_date=earned_date,
            credit_days=credit_days,
            reason=reason
        )

        return Response({
            "message": f"Comp off credited to {user.name}",
            "comp_off": CompOffSerializer(comp_off).data
        }, status=status.HTTP_201_CREATED)


class UseCompOffToReduceLOPView(APIView):
    """Use Comp Off to reduce existing LOP days"""

    def get(self, request):
        """Get LOP details that can be reduced"""
        from leaves.models import LeaveBalance, LeaveRequest

        user = request.user
        today = get_india_date()

        # Get available comp offs
        available_comp_offs = CompOff.objects.filter(
            user=user,
            status='earned',
            expires_on__gte=today
        ).order_by('expires_on')

        available_comp_off_days = sum(co.credit_days for co in available_comp_offs)

        # Get LOP from approved leave requests (current month and previous)
        lop_leaves = LeaveRequest.objects.filter(
            user=user,
            status='approved',
            lop_days__gt=0
        ).order_by('-start_date')[:10]  # Last 10 LOP leaves

        lop_details = [{
            'id': leave.id,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'total_days': float(leave.total_days),
            'lop_days': float(leave.lop_days),
            'paid_days': float(leave.paid_days),
            'comp_off_days': float(leave.comp_off_days),
            'leave_type': leave.leave_type.name if leave.leave_type else 'LOP',
            'reason': leave.reason
        } for leave in lop_leaves]

        # Format available comp offs for frontend
        available_comp_off_list = [{
            'id': co.id,
            'earned_date': co.earned_date,
            'credit_days': float(co.credit_days),
            'reason': co.reason,
            'expires_on': co.expires_on
        } for co in available_comp_offs]

        return Response({
            'available_comp_off_days': float(available_comp_off_days),
            'available_comp_offs': available_comp_off_list,
            'lop_leaves': lop_details
        })

    def post(self, request):
        """Use comp off to reduce LOP"""
        from leaves.models import LeaveRequest
        from accounts.models import Notification

        leave_request_id = request.data.get('leave_request_id')
        comp_off_id = request.data.get('comp_off_id')

        if not leave_request_id:
            return Response(
                {"error": "leave_request_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not comp_off_id:
            return Response(
                {"error": "comp_off_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get the leave request with LOP
        try:
            leave_request = LeaveRequest.objects.get(
                pk=leave_request_id,
                user=request.user,
                status='approved',
                lop_days__gt=0
            )
        except LeaveRequest.DoesNotExist:
            return Response(
                {"error": "Leave request not found or has no LOP"},
                status=status.HTTP_404_NOT_FOUND
            )

        today = get_india_date()

        # Get the selected comp off
        try:
            comp_off = CompOff.objects.get(
                pk=comp_off_id,
                user=request.user,
                status='earned',
                expires_on__gte=today
            )
        except CompOff.DoesNotExist:
            return Response(
                {"error": "Comp off not found or already used/expired"},
                status=status.HTTP_404_NOT_FOUND
            )

        days_to_reduce = float(comp_off.credit_days)

        # Check if days_to_reduce is valid - can only reduce up to available LOP
        actual_reduction = min(days_to_reduce, float(leave_request.lop_days))

        if actual_reduction < days_to_reduce:
            # Partial use - we'll use only what's needed
            remaining_days = days_to_reduce - actual_reduction

            # Mark original as used with reduced days
            comp_off.credit_days = actual_reduction
            comp_off.status = 'used'
            comp_off.used_date = today
            comp_off.save()

            # Create new comp off for remaining days
            CompOff.objects.create(
                user=request.user,
                earned_date=comp_off.earned_date,
                credit_days=remaining_days,
                status='earned',
                reason=f"Remaining after LOP reduction on {today}",
                expires_on=comp_off.expires_on
            )
        else:
            # Use entire comp off
            comp_off.status = 'used'
            comp_off.used_date = today
            comp_off.save()

        # Update leave request - reduce LOP, increase comp_off_days
        old_lop = leave_request.lop_days
        leave_request.lop_days = float(leave_request.lop_days) - actual_reduction
        leave_request.comp_off_days = float(leave_request.comp_off_days) + actual_reduction
        leave_request.save()

        # Create notification
        Notification.objects.create(
            user=request.user,
            title="LOP Reduced",
            message=f"Used {actual_reduction} Comp Off day(s) to reduce LOP. LOP reduced from {old_lop} to {leave_request.lop_days} days.",
            notification_type='leave_approved'
        )

        return Response({
            "message": f"Successfully reduced {actual_reduction} LOP day(s) using Comp Off",
            "leave_request": {
                "id": leave_request.id,
                "old_lop_days": float(old_lop),
                "new_lop_days": float(leave_request.lop_days),
                "comp_off_days_used": actual_reduction
            }
        })


class UseCompOffToCoverAbsentView(APIView):
    """Use Comp Off to cover absent days in attendance"""

    def get(self, request):
        """Get absent days and available comp-offs"""
        user = request.user
        today = get_india_date()

        # Get available comp offs (not expired)
        available_comp_offs = CompOff.objects.filter(
            user=user,
            status='earned',
            expires_on__gte=today
        ).order_by('expires_on')

        available_comp_off_days = sum(float(co.credit_days) for co in available_comp_offs)

        # Get absent attendance records (last 60 days)
        from datetime import timedelta
        sixty_days_ago = today - timedelta(days=60)

        absent_days = Attendance.objects.filter(
            user=user,
            status='absent',
            date__gte=sixty_days_ago,
            date__lte=today
        ).order_by('-date')

        absent_details = [{
            'id': att.id,
            'date': att.date,
            'working_hours': float(att.working_hours) if att.working_hours else 0,
            'notes': att.notes or ''
        } for att in absent_days]

        # Format available comp offs
        available_comp_off_list = [{
            'id': co.id,
            'earned_date': co.earned_date,
            'credit_days': float(co.credit_days),
            'reason': co.reason,
            'expires_on': co.expires_on
        } for co in available_comp_offs]

        return Response({
            'available_comp_off_days': available_comp_off_days,
            'available_comp_offs': available_comp_off_list,
            'absent_days': absent_details
        })

    def post(self, request):
        """Use comp off to cover an absent day"""
        from accounts.models import Notification

        attendance_id = request.data.get('attendance_id')
        comp_off_id = request.data.get('comp_off_id')

        if not attendance_id:
            return Response(
                {"error": "attendance_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not comp_off_id:
            return Response(
                {"error": "comp_off_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get the absent attendance record
        try:
            attendance = Attendance.objects.get(
                pk=attendance_id,
                user=request.user,
                status='absent'
            )
        except Attendance.DoesNotExist:
            return Response(
                {"error": "Absent attendance record not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        today = get_india_date()

        # Get the selected comp off
        try:
            comp_off = CompOff.objects.get(
                pk=comp_off_id,
                user=request.user,
                status='earned',
                expires_on__gte=today
            )
        except CompOff.DoesNotExist:
            return Response(
                {"error": "Comp off not found or already used/expired"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if comp-off has enough days (need at least 1 day for full absent, 0.5 for half)
        if float(comp_off.credit_days) < 1:
            # Half day comp-off - can only cover if we allow partial
            # For simplicity, we'll use the full comp-off value
            pass

        # Mark comp off as used
        comp_off.status = 'used'
        comp_off.used_date = today
        comp_off.save()

        # Update attendance - mark as present with comp-off note
        old_status = attendance.status
        attendance.status = 'present'
        attendance.working_hours = 8  # Full day
        attendance.notes = f"{attendance.notes}\n[Comp-Off Used: Earned on {comp_off.earned_date}]".strip()
        attendance.save(force_status=True)  # Prevent auto status calculation

        # Create notification
        Notification.objects.create(
            user=request.user,
            title="Absent Covered with Comp-Off",
            message=f"Used Comp-Off (earned on {comp_off.earned_date}) to cover absent on {attendance.date}. Status changed from Absent to Present.",
            notification_type='leave_approved'
        )

        return Response({
            "message": f"Successfully covered absent on {attendance.date} using Comp Off",
            "attendance": {
                "id": attendance.id,
                "date": attendance.date,
                "old_status": old_status,
                "new_status": attendance.status,
                "comp_off_used": {
                    "id": comp_off.id,
                    "earned_date": comp_off.earned_date,
                    "credit_days": float(comp_off.credit_days)
                }
            }
        })
